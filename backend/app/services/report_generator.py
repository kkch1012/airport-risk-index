"""
리포트 생성 서비스 — PDF, CSV, Excel
"""

import io
import logging
from datetime import datetime
from typing import Dict, List, Optional

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.risk_history import CategoryScoreRecord, RiskAssessment

logger = logging.getLogger(__name__)

CATEGORY_LABELS = {
    "weather": "기상위험",
    "aviation": "항공안전",
    "security": "보안위협",
    "health": "보건위험",
    "operational": "운영위험",
    "external": "외부요인",
}

LEVEL_LABELS = {
    "LOW": "낮음",
    "MODERATE": "보통",
    "HIGH": "높음",
    "CRITICAL": "심각",
}


class ReportGenerator:
    """위험지수 리포트 생성"""

    def __init__(self, session: AsyncSession):
        self.session = session

    async def _fetch_data(
        self,
        airport_code: Optional[str] = None,
    ) -> List[dict]:
        """리포트용 데이터 조회"""
        stmt = (
            select(RiskAssessment)
            .order_by(RiskAssessment.assessed_date.desc())
            .limit(100)
        )
        if airport_code:
            stmt = stmt.where(RiskAssessment.airport_code == airport_code)

        result = await self.session.execute(stmt)
        assessments = result.scalars().all()

        if not assessments:
            return []

        # N+1 방지: 모든 assessment_id에 대한 카테고리 점수를 한 번에 조회
        assessment_ids = [a.id for a in assessments]
        cat_stmt = select(CategoryScoreRecord).where(
            CategoryScoreRecord.assessment_id.in_(assessment_ids)
        )
        cat_result = await self.session.execute(cat_stmt)
        all_cats = cat_result.scalars().all()

        # assessment_id별로 그룹핑
        cats_by_assessment: Dict[int, Dict[str, float]] = {}
        proxy_by_assessment: Dict[int, List[str]] = {}
        nodata_by_assessment: Dict[int, List[str]] = {}
        for c in all_cats:
            cats_by_assessment.setdefault(c.assessment_id, {})[c.category_code] = c.score
            # 구 레코드(컬럼 추가 전)는 getattr 기본값으로 안전 처리
            if getattr(c, "is_proxy", False):
                proxy_by_assessment.setdefault(c.assessment_id, []).append(c.category_code)
            if getattr(c, "has_data", True) is False:
                nodata_by_assessment.setdefault(c.assessment_id, []).append(c.category_code)

        rows = []
        for a in assessments:
            cats = cats_by_assessment.get(a.id, {})
            rows.append({
                "airport_code": a.airport_code,
                "airport_name": a.airport_name,
                "date": str(a.assessed_date),
                "total_score": a.total_score,
                "risk_level": a.risk_level,
                "confidence_note": self._confidence_note(
                    proxy_by_assessment.get(a.id, []),
                    nodata_by_assessment.get(a.id, []),
                ),
                **{f"cat_{k}": v for k, v in cats.items()},
            })

        return rows

    @staticmethod
    def _confidence_note(proxy_codes: List[str], nodata_codes: List[str]) -> str:
        """카테고리별 신뢰도 비고 문자열 생성.

        추정치(뉴스 프록시)·데이터없음 카테고리를 사람이 읽을 수 있게 표기한다.
        """
        parts = []
        if proxy_codes:
            parts.append(
                "추정: " + ", ".join(CATEGORY_LABELS.get(c, c) for c in proxy_codes)
            )
        if nodata_codes:
            parts.append(
                "데이터없음: " + ", ".join(CATEGORY_LABELS.get(c, c) for c in nodata_codes)
            )
        return " / ".join(parts)

    # ── CSV ────────────────────────────────────

    async def generate_csv(
        self, airport_code: Optional[str] = None
    ) -> bytes:
        """CSV 리포트 생성"""
        import pandas as pd

        rows = await self._fetch_data(airport_code)
        if not rows:
            rows = [{"message": "데이터 없음"}]

        df = pd.DataFrame(rows)

        # 컬럼명 한글화
        rename_map = {
            "airport_code": "공항코드",
            "airport_name": "공항명",
            "date": "평가일",
            "total_score": "종합점수",
            "risk_level": "위험등급",
            "confidence_note": "신뢰도비고",
        }
        for code, label in CATEGORY_LABELS.items():
            rename_map[f"cat_{code}"] = label

        df = df.rename(columns=rename_map)

        buf = io.BytesIO()
        df.to_csv(buf, index=False, encoding="utf-8-sig")
        return buf.getvalue()

    # ── Excel ──────────────────────────────────

    async def generate_excel(
        self, airport_code: Optional[str] = None
    ) -> bytes:
        """Excel 리포트 생성"""
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment

        rows = await self._fetch_data(airport_code)
        if not rows:
            rows = [{"message": "데이터 없음"}]

        df = pd.DataFrame(rows)

        rename_map = {
            "airport_code": "공항코드",
            "airport_name": "공항명",
            "date": "평가일",
            "total_score": "종합점수",
            "risk_level": "위험등급",
            "confidence_note": "신뢰도비고",
        }
        for code, label in CATEGORY_LABELS.items():
            rename_map[f"cat_{code}"] = label

        df = df.rename(columns=rename_map)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="위험지수")

            ws = writer.sheets["위험지수"]

            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # 컬럼 너비 자동 조정
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        return buf.getvalue()

    # ── PDF ────────────────────────────────────

    async def generate_pdf(
        self, airport_code: Optional[str] = None
    ) -> bytes:
        """PDF 리포트 생성"""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        rows = await self._fetch_data(airport_code)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)

        # 폰트 등록 (시스템 폰트 없으면 기본 Helvetica)
        font_name = "Helvetica"
        try:
            pdfmetrics.registerFont(TTFont("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"))
            font_name = "NanumGothic"
        except Exception:
            try:
                pdfmetrics.registerFont(TTFont("AppleGothic", "/System/Library/Fonts/AppleSDGothicNeo.ttc"))
                font_name = "AppleGothic"
            except Exception:
                pass

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Title"],
            fontName=font_name, fontSize=16,
        )
        normal_style = ParagraphStyle(
            "CustomNormal", parent=styles["Normal"],
            fontName=font_name, fontSize=10,
        )
        cell_style = ParagraphStyle(
            "CustomCell", parent=styles["Normal"],
            fontName=font_name, fontSize=8,
        )

        elements = []

        # 제목
        title = "Airport Risk Index Report"
        if airport_code:
            title += f" - {airport_code}"
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 10 * mm))

        generated = datetime.now().strftime("%Y-%m-%d %H:%M")
        elements.append(Paragraph(f"Generated: {generated}", normal_style))
        elements.append(Spacer(1, 5 * mm))

        if not rows:
            elements.append(Paragraph("No data available", normal_style))
        else:
            # 테이블 데이터
            headers = ["Airport", "Date", "Score", "Level", "비고(신뢰도)"]
            table_data = [headers]
            for r in rows[:50]:  # 최대 50행
                level = LEVEL_LABELS.get(r["risk_level"], r["risk_level"])
                note = r.get("confidence_note", "") or "-"
                table_data.append([
                    f'{r["airport_code"]}',
                    r["date"],
                    f'{r["total_score"]:.1f}',
                    level,
                    Paragraph(note, cell_style),  # 줄바꿈 위해 Paragraph 사용
                ])

            table = Table(table_data, colWidths=[60, 70, 45, 50, 180])

            # 위험등급별 색상
            style_commands = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]

            # HIGH/CRITICAL 행 강조
            for i, r in enumerate(rows[:50], start=1):
                if r["risk_level"] == "CRITICAL":
                    style_commands.append(
                        ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#fef2f2"))
                    )
                elif r["risk_level"] == "HIGH":
                    style_commands.append(
                        ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#fffbeb"))
                    )

            table.setStyle(TableStyle(style_commands))
            elements.append(table)

        doc.build(elements)
        return buf.getvalue()
